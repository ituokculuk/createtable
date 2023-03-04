from flask import Flask, redirect, render_template, request, Response, url_for,send_file
from werkzeug.utils import secure_filename
import os,sys
import subprocess
from openpyxl import load_workbook
from operator import attrgetter
import re
import openpyxl
from openpyxl.styles import PatternFill
import os.path
from flask_frozen import Freezer

# sayfayı yebilerse her eşyi baştan yapsıııınn dosyalar kalıyo

app = Flask(__name__)
freezer = Freezer(app)

@app.route("/", methods = ['GET', 'POST'])
def main():
    return render_template("homepage.html")


@app.route("/upload", methods = ['GET', 'POST'])
def upload():
    return render_template("uploadpage.html",sheet=secure_filename(""))


## file uploading stuff
@app.route("/getfile", methods = ['GET','POST'])
def getfile():
    global file
    if request.method == 'POST': 
        file = request.files['file']
        filename_delete = file.filename
        file.save(os.path.join(os.path.dirname(__file__),secure_filename(file.filename)))
        createTablo(secure_filename(file.filename))
        if os.path.isfile(secure_filename(file.filename)):
            os.remove(secure_filename(file.filename))
    return render_template('uploadpage.html',sheet=secure_filename("table.xlsx"))



@app.route("/download")
def download_file():
    return send_file(secure_filename("table.xlsx"), as_attachment=True)


def createTablo(file):
    wb = load_workbook(file)
    #ws = wb.sheet_by_index(0)
    #ws = wb.active(0)
    ws = wb.worksheets[0]

    if not os.path.isfile("table.xlsx"):
        workbook = openpyxl.Workbook("table.xlsx")
        workbook.save("table.xlsx")
        
        
    workbook = load_workbook("table.xlsx", data_only = True) # Workbook() takes one, non-optional, argument  which is the filename that we want to create.
    worksheet = workbook.worksheets[0] # The workbook object is then used to add new worksheet via the add_worksheet() method.
    workbook.save("table.xlsx")

    for i in range(ws.max_row ):
        for j in range(ws.max_column):
            worksheet.cell(i+1,j+1).value = None

    isim_column = 0
    seans_column = 0
    student_number_in_seans = 50 # gunluk ogrenci kisitlamasi bunu kullanıcıdan al

    student_list = [] 
    seans_list = []
    Seans_array = []
    days_array = ["pazartesi","salı","çarşamba","perşembe","cuma","cumartesi","pazar"]
    color_list = []
    colors=["C5FFCC","FFFF88","FFA0CC","CCFFFF","969696","FFFFFF"]
    color_index = 0   

    #------------------- CLASSES ------------------------
    class Students:
        color = ""
        name = ""
        seans = ""
        numberofseans = 0
        yerlestirildi = False
        
        
        def __init__(self, name, seans, color):
            self.name = name
            self.seans = seans
            self.color = color
            
    class Seans:
        day = ""
        minute = ""
        seans = ""
        total = 0
        
        def __init__(self, day, minute,seans,total):
            self.day = day
            self.minute = minute   
            self.seans = seans   
            self.total = total 



    #---------------------------------------------------
    # isim ve gunlerin satirini belirler
    for i in range(1):
        for j in range(ws.max_column):
            if(ws.cell(i+1,j+1).value != None):
                if("isim" in ws.cell(i+1,j+1).value or "ad" in ws.cell(i+1,j+1).value):
                    isim_column = j
                if("seans" in ws.cell(i+1,j+1).value or "Seans" in ws.cell(i+1,j+1).value):
                    seans_column = j

      
    # ogrencilerin isimilerini ve sectigi gunleri cekme
    for i in range(0,ws.max_row):
        if( ws.cell(i+1,isim_column+1).value == None or "isim" in ws.cell(i+1,isim_column+1).value or "ad" in ws.cell(i+1,isim_column+1).value):
            pass
        else:
            student_list.append(Students(ws.cell(i+1,isim_column+1).value,ws.cell(i+1,seans_column+1).value,ws.cell(i+1,seans_column+1).fill.start_color.index))
            
            if(color_list.count(ws.cell(i+1,seans_column+1).fill.start_color.index) == 0):
                color_list.append(ws.cell(i+1,seans_column+1).fill.start_color.index)

    # gunleri ve seanslari belirleme
    for i in range(len(student_list)):
        #ogrencinin sectigi seanslarin arrayini olusturur
        student_list[i].seans=student_list[i].seans.replace(" ", "")
        parsedSeans = student_list[i].seans.split(",")
        student_list[i].numberofseans = len(parsedSeans)
        for j in range(len(parsedSeans)):
            if(seans_list.count(parsedSeans[j]) == 0 and "" not in parsedSeans):
                seans_list.append(parsedSeans[j])
                
    # seans list günleri sıralama
    for i in range(len(seans_list)):
        my_list = re.split(r'(\d+)', seans_list[i])
        day = days_array.index(my_list[0].lower())
        minute = (((int)(my_list[1]))*60 + (int)(my_list[3]))
        Seans_array.append(Seans(day, minute,seans_list[i],(day*10000+minute)))
        Seans_array.sort(key=attrgetter("total"))

    for i in range(len(Seans_array)):
        seans_list[i] = Seans_array[i].seans

    # baslık
    for i in range(len(seans_list)):
        worksheet.cell(row=0+1, column=i+1).value = seans_list[i]


    #siralama
    student_list.sort(key=attrgetter("numberofseans"))

    """
    for i in range(len(student_list)):
        print(student_list[i].name,student_list[i].seans,student_list[i].numberofseans)
    """



    for colornumber in range(len(color_list)):
        # yerlestirme len(student_list)
        for i in range(len(student_list)):
            if(student_list[i].yerlestirildi == False and color_list[colornumber] == student_list[i].color):
                parsedSeans = student_list[i].seans.split(",")
                maxplace=0
                RowIndex_seans=0
                ColumnIndex_seans=0
                maxColumn_Index = 0
                maxRos_Index = 0
                
                for k in range(len(parsedSeans)):
                    ColumnIndex_seans = seans_list.index(parsedSeans[k])
                    for j in range(student_number_in_seans):
                        if(worksheet.cell(j+1, ColumnIndex_seans+1).value == None):
                            RowIndex_seans = j
                            break  
                    if(student_number_in_seans-RowIndex_seans > maxplace):
                        maxplace = student_number_in_seans-RowIndex_seans
                        maxColumn_Index = ColumnIndex_seans 
                        maxRos_Index = RowIndex_seans
                    
                worksheet.cell(maxRos_Index+1,maxColumn_Index+1).value = student_list[i].name
                if(color_index<len(colors)):
                    worksheet.cell(maxRos_Index+1,maxColumn_Index+1).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb=colors[color_index]))
                else: 
                    worksheet.cell(maxRos_Index+1,maxColumn_Index+1).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb="FFFFFF"))
    
                student_list[i].yerlestirildi = True
        color_index=color_index+1        
                    
                    
                
    workbook.save("table.xlsx")
    workbook.close()
    

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "build":
        freezer.freeze()
    else:
        app.run(debug=True)